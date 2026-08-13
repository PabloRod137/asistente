import os
import re
import sys
import unicodedata
import logging
import sqlite3
from datetime import datetime
import openpyxl

# Configurar logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger("importador_clientes")

# Añadir el directorio raíz al path para poder importar database.py
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import database

# Firma de la fila de ejemplo de la plantilla en blanco (plantilla_clientes_maira.xlsx). Se
# detecta por CONTENIDO, no por posición fija: algunos clientes (como LexGuardian) sobrescriben
# directamente esa fila con su primer cliente real en vez de insertar una fila nueva debajo.
FILA_EJEMPLO_NOMBRE = "ana lopez garcia"
FILA_EJEMPLO_NIF = "12345678a"

# Patrón para extraer un número de expediente tipo AAAA-NNN de una ruta de SharePoint, cuando
# la columna "Número de expediente" viene vacía pero el propio nombre de la carpeta ya lo lleva
# (ej. ".../2013-001 - Autónomo - BERDEJO PAYNO, ALBERTO/00_FACTURAS/").
PATRON_EXPEDIENTE_EN_RUTA = re.compile(r'(\d{4}-\d{3,4})\s*-')


def normalizar_cabecera(val):
    if val is None:
        return ""
    s = str(val).strip().lower()
    s = "".join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s


def parsear_fecha(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def es_fila_de_ejemplo(nombre, nif_cif) -> bool:
    nombre_norm = normalizar_cabecera(nombre) if nombre else ""
    nif_norm = normalizar_cabecera(nif_cif) if nif_cif else ""
    return nombre_norm == FILA_EJEMPLO_NOMBRE and nif_norm == FILA_EJEMPLO_NIF


def extraer_expediente_de_ruta(carpeta_sharepoint: str):
    if not carpeta_sharepoint:
        return None
    m = PATRON_EXPEDIENTE_EN_RUTA.search(carpeta_sharepoint)
    return m.group(1) if m else None


def importar(excel_path):
    if not os.path.exists(excel_path):
        print(f"Error: El archivo '{excel_path}' no existe.")
        sys.exit(1)

    print(f"Cargando archivo Excel: {excel_path}...")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"Error al abrir el archivo Excel: {e}")
        sys.exit(1)

    if "Clientes" not in wb.sheetnames:
        print("Error: El archivo Excel no contiene la hoja obligatoria 'Clientes'.")
        sys.exit(1)

    sheet = wb["Clientes"]
    if sheet.max_row < 1:
        print("Error: La hoja 'Clientes' está vacía.")
        sys.exit(1)

    # 1. Mapear cabeceras a índices
    cabeceras_fila = [cell.value for cell in sheet[1]]
    cabeceras_mapeadas = {}
    for idx, val in enumerate(cabeceras_fila):
        norm = normalizar_cabecera(val)
        if norm:
            cabeceras_mapeadas[norm] = idx + 1  # 1-based index para openpyxl

    # 2. Validar columnas obligatorias — solo el nombre es realmente imprescindible. El
    # teléfono decide a qué tabla va la fila (clientes reales vs lista de espera), y NIF /
    # expediente son útiles pero no bloquean el alta: Maira solo necesita nombre + teléfono
    # para reconocer a alguien; el resto lo puede ir completando después.
    if "nombre completo" not in cabeceras_mapeadas:
        print("Error: Falta la columna obligatoria 'Nombre completo' en el Excel.")
        sys.exit(1)
    if "telefono (whatsapp)" not in cabeceras_mapeadas:
        print("Error: Falta la columna 'Teléfono (WhatsApp)' en el Excel (aunque esté vacía en algunas filas, la columna debe existir).")
        sys.exit(1)

    # Inicializar contadores e informe
    procesados = 0
    actualizados = 0
    pendientes_cargados = 0
    omitidos = 0
    errores = []

    conn = database.get_connection()
    cursor = conn.cursor()

    # Mapeo de columnas opcionales
    opcionales = {
        "tipo_cliente": cabeceras_mapeadas.get("tipo de cliente"),
        "email": cabeceras_mapeadas.get("email"),
        "empresa": cabeceras_mapeadas.get("empresa"),
        "carpeta_sharepoint": cabeceras_mapeadas.get("carpeta sharepoint"),
        "gestor_asignado": cabeceras_mapeadas.get("gestor asignado"),
        "idioma_preferido": cabeceras_mapeadas.get("idioma preferido"),
        "fecha_alta": cabeceras_mapeadas.get("fecha de alta"),
        "notas": cabeceras_mapeadas.get("notas"),
    }
    col_nif = cabeceras_mapeadas.get("nif/cif")
    col_expediente = cabeceras_mapeadas.get("numero de expediente")

    print("Iniciando procesamiento de filas (saltando cabecera y, si aparece, la fila de ejemplo)...")

    for r_idx in range(2, sheet.max_row + 1):
        row_cells = [sheet.cell(row=r_idx, column=c_idx).value for c_idx in range(1, len(cabeceras_fila) + 1)]
        if all(val is None or str(val).strip() == "" for val in row_cells):
            continue  # Fila vacía, saltar silenciosamente

        nombre = sheet.cell(row=r_idx, column=cabeceras_mapeadas["nombre completo"]).value
        telefono_raw = sheet.cell(row=r_idx, column=cabeceras_mapeadas["telefono (whatsapp)"]).value
        nif_cif_raw = sheet.cell(row=r_idx, column=col_nif).value if col_nif else None
        num_expediente_raw = sheet.cell(row=r_idx, column=col_expediente).value if col_expediente else None

        if es_fila_de_ejemplo(nombre, nif_cif_raw):
            print(f"ℹ️ Fila {r_idx}: detectada como la fila de ejemplo de la plantilla, se omite.")
            continue

        if not nombre or str(nombre).strip() == "":
            omitidos += 1
            err_msg = f"Fila {r_idx}: Saltada, Nombre completo vacío"
            print(f"⚠️ {err_msg}")
            errores.append(err_msg)
            continue
        nombre = str(nombre).strip()

        nif_cif = str(nif_cif_raw).strip().upper() if nif_cif_raw and str(nif_cif_raw).strip() else None

        # Extraer campos opcionales
        def read_opt(key, default=None):
            col_idx = opcionales[key]
            if col_idx is None:
                return default
            val = sheet.cell(row=r_idx, column=col_idx).value
            return str(val).strip() if val is not None and str(val).strip() else default

        # La columna "Tipo de cliente" está pensada para el estado en el CRM ('nuevo'/'activo'),
        # pero es habitual que se rellene con la forma legal del cliente (Autónomo/Empresa/Otros)
        # en vez de eso. Si el valor no es un estado CRM válido, se guarda como información en
        # las notas y el cliente se marca 'activo' por defecto (viene de un listado real de
        # clientes ya existentes del despacho, no son leads "nuevos").
        tipo_cliente_raw = read_opt("tipo_cliente")
        if tipo_cliente_raw and tipo_cliente_raw.strip().lower() in ("nuevo", "activo"):
            tipo_cliente = tipo_cliente_raw.strip().lower()
            forma_legal = None
        else:
            tipo_cliente = "activo"
            forma_legal = tipo_cliente_raw

        email = read_opt("email")
        empresa = read_opt("empresa")
        carpeta_sharepoint = read_opt("carpeta_sharepoint")
        gestor_asignado = read_opt("gestor_asignado")
        idioma_preferido = read_opt("idioma_preferido", "es")
        fecha_alta = parsear_fecha(sheet.cell(row=r_idx, column=opcionales["fecha_alta"]).value if opcionales["fecha_alta"] else None)
        notas = read_opt("notas")
        if forma_legal:
            notas = f"{notas} | Tipo: {forma_legal}" if notas else f"Tipo: {forma_legal}"

        num_expediente = str(num_expediente_raw).strip() if num_expediente_raw and str(num_expediente_raw).strip() else None
        if not num_expediente:
            extraido = extraer_expediente_de_ruta(carpeta_sharepoint)
            if extraido:
                num_expediente = extraido

        telefono_norm = database.normalizar_telefono(str(telefono_raw).strip()) if telefono_raw and str(telefono_raw).strip() else None

        try:
            if telefono_norm:
                # Camino normal: hay teléfono, va directo a la tabla de clientes reales.
                now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cursor.execute("SELECT phone_number FROM clientes WHERE phone_number = ?", (telefono_norm,))
                existe = cursor.fetchone() is not None

                if existe:
                    cursor.execute("""
                        UPDATE clientes
                        SET nombre = ?, empresa = ?, email = ?, notas = ?, ultima_visita = ?,
                            numero_expediente = ?, tipo_cliente = ?, nif_cif = ?, fecha_alta = COALESCE(?, fecha_alta),
                            gestor_asignado = ?, carpeta_sharepoint = ?, idioma_preferido = ?
                        WHERE phone_number = ?
                    """, (
                        nombre, empresa, email, notas, now_str, num_expediente,
                        tipo_cliente, nif_cif, fecha_alta, gestor_asignado,
                        carpeta_sharepoint, idioma_preferido, telefono_norm
                    ))
                    actualizados += 1
                    print(f"🔄 Fila {r_idx}: Cliente actualizado ({telefono_norm} - {nombre})")
                else:
                    cursor.execute("""
                        INSERT INTO clientes (
                            phone_number, nombre, empresa, email, notas, primera_visita, ultima_visita,
                            total_conversaciones, numero_expediente, tipo_cliente, nif_cif, fecha_alta,
                            gestor_asignado, carpeta_sharepoint, idioma_preferido
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        telefono_norm, nombre, empresa, email, notas, now_str, now_str,
                        num_expediente, tipo_cliente, nif_cif, fecha_alta or now_str[:10],
                        gestor_asignado, carpeta_sharepoint, idioma_preferido
                    ))
                    procesados += 1
                    print(f"📥 Fila {r_idx}: Cliente insertado ({telefono_norm} - {nombre})")
                conn.commit()
            else:
                # Sin teléfono: a la lista de espera. Maira lo enlazará sola la primera vez
                # que esa persona escriba y confirme que es quien la ficha dice que es.
                database.insertar_cliente_pendiente(
                    nombre=nombre, nif_cif=nif_cif, numero_expediente=num_expediente,
                    notas=notas, carpeta_sharepoint=carpeta_sharepoint,
                    tipo_cliente=tipo_cliente or "activo", idioma_preferido=idioma_preferido or "es"
                )
                pendientes_cargados += 1
                print(f"⏳ Fila {r_idx}: Sin teléfono, cargado en lista de espera ({nombre})")

        except sqlite3.Error as db_err:
            conn.rollback()
            omitidos += 1
            err_msg = f"Fila {r_idx}: Error de base de datos al guardar ({db_err})"
            print(f"❌ {err_msg}")
            errores.append(err_msg)

    conn.close()

    print("\n" + "="*60)
    print("           RESUMEN DE LA IMPORTACIÓN")
    print("="*60)
    print(f"Clientes nuevos insertados (con teléfono): {procesados}")
    print(f"Clientes existentes actualizados: {actualizados}")
    print(f"Cargados en lista de espera (sin teléfono todavía): {pendientes_cargados}")
    print(f"Filas omitidas / con errores: {omitidos}")
    print(f"Total filas de datos procesadas: {procesados + actualizados + pendientes_cargados + omitidos}")
    print("="*60)

    if errores:
        print("\nDetalle de filas omitidas/errores:")
        for err in errores:
            print(f" - {err}")
    print("="*60 + "\n")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python scripts/importar_clientes.py {ruta_archivo_excel}")
        sys.exit(1)
    importar(sys.argv[1])
