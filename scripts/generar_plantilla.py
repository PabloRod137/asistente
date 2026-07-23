import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def generar_plantilla(filepath="plantilla_clientes_maira.xlsx"):
    wb = openpyxl.Workbook()
    
    # Hoja 1: Instrucciones
    ws_inst = wb.active
    ws_inst.title = "Instrucciones"
    ws_inst.views.sheetView[0].showGridLines = True
    
    ws_inst["A1"] = "Instrucciones de uso del Importador de Clientes Maira"
    ws_inst["A1"].font = Font(name="Arial", size=16, bold=True)
    
    ws_inst["A3"] = "1. Rellena la hoja 'Clientes' con los datos reales de tu base de datos."
    ws_inst["A4"] = "2. Las siguientes columnas son OBLIGATORIAS y no pueden estar vacías:"
    ws_inst["A5"] = "   - Nombre completo"
    ws_inst["A6"] = "   - Teléfono (WhatsApp)"
    ws_inst["A7"] = "   - NIF/CIF"
    ws_inst["A8"] = "   - Número de expediente"
    ws_inst["A9"] = "3. La fila 2 contiene un ejemplo ficticio sombreado que el importador omitirá automáticamente."
    ws_inst["A10"] = "4. Puedes cambiar el orden de las columnas si lo necesitas; el importador las busca por su nombre exacto en la fila 1."
    
    # Hoja 2: Clientes
    ws_cli = wb.create_sheet(title="Clientes")
    ws_cli.views.sheetView[0].showGridLines = True
    
    headers = [
        "Nombre completo", "Teléfono (WhatsApp)", "NIF/CIF", "Número de expediente",
        "Tipo de cliente", "Email", "Empresa", "Carpeta SharePoint",
        "Gestor asignado", "Idioma preferido", "Fecha de alta", "Notas"
    ]
    
    # Fila 1: Cabeceras
    for col_idx, text in enumerate(headers, 1):
        cell = ws_cli.cell(row=1, column=col_idx, value=text)
        cell.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Fila 2: Ejemplo
    ejemplo = [
        "Ana López García", "34600112233", "12345678A", "EXP-2026-001",
        "activo", "ana.lopez@ejemplo.com", "Panadería Ana S.L.", "/Clientes/Ana Lopez Garcia/",
        "Alberto Berdejo", "es", "2026-01-15", "Cliente desde apertura del negocio, prefiere WhatsApp"
    ]
    for col_idx, text in enumerate(ejemplo, 1):
        cell = ws_cli.cell(row=2, column=col_idx, value=text)
        cell.font = Font(name="Arial", size=10, italic=True, color="595959")
        cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        
    wb.save(filepath)
    print(f"Plantilla generada con éxito en '{filepath}'")

if __name__ == "__main__":
    generar_plantilla()
