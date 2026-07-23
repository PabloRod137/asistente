import os
import sys
import sqlite3
import asyncio
import logging
import openpyxl

# Configurar stdout a UTF-8 para emojis en Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
os.environ["MODO_TEST"] = "true"

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger("test_fase6")

def check_db_columns():
    import database
    database.init_db()
    conn = database.get_connection()
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(clientes)")
    columns = [row[1] for row in cursor.fetchall()]
    conn.close()
    
    assert "carpeta_sharepoint" in columns, "Columna carpeta_sharepoint no encontrada"
    assert "idioma_preferido" in columns, "Columna idioma_preferido no encontrada"
    logger.info("✅ Columnas carpeta_sharepoint e idioma_preferido verificadas en SQLite.")

def crear_excel_test_normal(filename="test_normal.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    headers = [
        "Nombre completo", "Teléfono (WhatsApp)", "NIF/CIF", "Número de expediente",
        "Tipo de cliente", "Email", "Empresa", "Carpeta SharePoint",
        "Gestor asignado", "Idioma preferido", "Fecha de alta", "Notas"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        
    # Fila 2: Ejemplo (omitida)
    ws.append(["Ejemplo Nombre", "34600000000", "00000000T", "EXP-EX", "activo", "ex@mail.com", "Emp", "/path/", "Gestor", "es", "2026-01-01", "Notas"])
    
    # Fila 3: Válido 1
    ws.append(["Juan Pérez", "34600000001", "11111111A", "EXP-001", "activo", "juan@mail.com", "Pérez S.L.", "/SP/Juan", "Alberto", "es", "2026-02-01", "Notas Juan"])
    
    # Fila 4: Inválido (NIF vacío)
    ws.append(["Pedro Gómez", "34600000002", "", "EXP-002", "activo", "pedro@mail.com", "Gomez S.L.", "", "", "es", "", ""])
    
    # Fila 5: Upsert sobre Válido 1 (Mismo teléfono, nombre modificado)
    ws.append(["Juan Pérez Modificado", "34600000001", "11111111A", "EXP-001-MOD", "activo", "juan@mail.com", "Pérez S.L.", "/SP/Juan", "Alberto", "es", "2026-02-01", "Notas Actualizadas"])
    
    wb.save(filename)
    logger.info(f"✅ Excel de test normal '{filename}' generado.")

def crear_excel_test_reordenado(filename="test_reordenado.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Columnas reordenadas a propósito
    headers_reordered = [
        "NIF/CIF", "Número de expediente", "Teléfono (WhatsApp)", "Nombre completo",
        "Carpeta SharePoint", "Idioma preferido", "Fecha de alta"
    ]
    for col, h in enumerate(headers_reordered, 1):
        ws.cell(row=1, column=col, value=h)
        
    # Fila 2: Ejemplo (omitida)
    ws.append(["00000000T", "EXP-EX", "34600000000", "Ejemplo Nombre", "/path/", "es", "2026-01-01"])
    
    # Fila 3: Válido Reordenado
    ws.append(["22222222B", "EXP-003", "34600000003", "María López", "/SP/Maria", "en", "2026-03-01"])
    
    wb.save(filename)
    logger.info(f"✅ Excel de test reordenado '{filename}' generado.")

async def test_maira_conversation():
    import main
    import database
    
    # Simular mensaje de un cliente importado ("María López" con teléfono 34600000003)
    # p_reg = "34600000003"
    # Este cliente ya está importado con tipo_cliente = 'activo' y nombre = 'María López'
    # Debería saludar por su nombre y no disparar el alta
    
    phone = "34600000003"
    
    # Verificar primero en DB que está activo
    cli = database.get_cliente_by_phone(phone)
    assert cli is not None, "El cliente María López debe existir en la base de datos"
    assert cli["nombre"] == "María López", f"El nombre debe ser María López, obtenido: {cli['nombre']}"
    assert cli["tipo_cliente"] == "activo", f"El cliente debe estar activo, obtenido: {cli['tipo_cliente']}"
    
    logger.info("Simulando conversación de WhatsApp de cliente importado...")
    res = await main.procesar_flujo_mensaje(phone, "hola, buenas, quería pedir una cita", "text")
    logger.info(f"Respuesta del bot:\n\"{res}\"\n")
    
    # Comprobar que responde saludando o reconociendo al cliente sin flujo de alta
    assert res is not None, "La respuesta no puede ser None"
    assert "nombre" not in res.lower(), "El bot no debería solicitar el nombre porque ya está registrado"
    assert "nif" not in res.lower(), "El bot no debería solicitar el NIF"
    logger.info("✅ Integración con el reconocimiento de clientes (Fase 3) verificada correctamente.")

def cleanup_files():
    for f in ["test_normal.xlsx", "test_reordenado.xlsx"]:
        if os.path.exists(f):
            os.remove(f)

async def main_test():
    import database
    from scripts import importar_clientes
    
    # Limpiar base de datos para la prueba
    conn = database.get_connection()
    c = conn.cursor()
    c.execute("DELETE FROM clientes WHERE phone_number IN ('34600000001', '34600000002', '34600000003')")
    conn.commit()
    conn.close()

    print("\n" + "="*80)
    print("             EJECUCIÓN DE VERIFICACIONES DE FASE 6")
    print("="*80 + "\n")

    # 1. Verificar esquema de DB
    check_db_columns()
    
    # Generar archivos de prueba
    crear_excel_test_normal()
    crear_excel_test_reordenado()

    # 2. Ejecutar importación normal
    print("\n--- Ejecutando importación con columnas estándar ---")
    importar_clientes.importar("test_normal.xlsx")
    
    # Verificar resultados en SQLite
    cli_1 = database.get_cliente_by_phone("34600000001")
    assert cli_1 is not None, "Juan Pérez no fue importado"
    # Debe ser la versión de Fila 5 debido al Upsert
    assert cli_1["nombre"] == "Juan Pérez Modificado", "El upsert por teléfono no funcionó"
    assert cli_1["numero_expediente"] == "EXP-001-MOD", "El número de expediente no se actualizó"
    assert cli_1["tipo_cliente"] == "activo", "El tipo de cliente debe ser 'activo'"
    assert cli_1["carpeta_sharepoint"] == "/SP/Juan", "Carpeta SharePoint no guardada"
    assert cli_1["idioma_preferido"] == "es", "Idioma preferido incorrecto"
    
    cli_2 = database.get_cliente_by_phone("34600000002")
    assert cli_2 is None, "Pedro Gómez (NIF vacío) fue importado incorrectamente"
    logger.info("✅ Validaciones de obligatorios y Upsert verificados en SQLite.")

    # 3. Ejecutar importación del reordenado (mapeo por nombre de columna)
    print("\n--- Ejecutando importación con columnas reordenadas ---")
    importar_clientes.importar("test_reordenado.xlsx")
    
    cli_3 = database.get_cliente_by_phone("34600000003")
    assert cli_3 is not None, "María López (Excel reordenado) no fue importada"
    assert cli_3["nombre"] == "María López", "Mapeo incorrecto de Nombre en columnas reordenadas"
    assert cli_3["nif_cif"] == "22222222B", "Mapeo incorrecto de NIF en columnas reordenadas"
    assert cli_3["carpeta_sharepoint"] == "/SP/Maria", "Mapeo incorrecto de SharePoint en columnas reordenadas"
    assert cli_3["idioma_preferido"] == "en", "Mapeo incorrecto de Idioma preferido"
    logger.info("✅ Mapeo dinámico por nombre de columna verificado con éxito.")

    # 4. Segunda ejecución del importador (no duplicación)
    print("\n--- Ejecutando importación repetida del reordenado ---")
    importar_clientes.importar("test_reordenado.xlsx")
    conn = database.get_connection()
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM clientes WHERE phone_number = '34600000003'")
    count = c.fetchone()[0]
    conn.close()
    assert count == 1, f"Se detectaron registros duplicados para María López (Count: {count})"
    logger.info("✅ Segunda ejecución no duplicó registros (Upsert verificado).")

    # 5. Flujo de conversación (Integración con Fase 3)
    await test_maira_conversation()

    # Limpieza
    cleanup_files()
    
    print("\n" + "="*80)
    print("🏆 ¡TODAS LAS VERIFICACIONES DE LA FASE 6 PASADAS CORRECTAMENTE!")
    print("="*80 + "\n")

if __name__ == "__main__":
    asyncio.run(main_test())
